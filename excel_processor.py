"""
Excel processor: scans worksheets, masks PII, and preserves workbook formatting as much as openpyxl allows.
Only .xlsx is supported to avoid destructive format conversion.
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from pii_engine import mask_value, should_mask

MASKED_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")


def process_excel(input_path: str, output_path: str, highlight_masked: bool = True) -> dict:
    report = {
        "file": os.path.basename(input_path),
        "sheets_processed": 0,
        "total_cells_scanned": 0,
        "total_pii_masked": 0,
        "pii_breakdown": {},
        "sheet_details": [],
        "errors": [],
    }

    try:
        wb = load_workbook(input_path)
    except Exception as e:
        report["errors"].append(f"Could not open workbook: {e}")
        return report

    for ws in wb.worksheets:
        sheet_info = {
            "sheet": ws.title,
            "cells_scanned": 0,
            "pii_found": 0,
            "columns_masked": [],
        }

        header_row = None
        headers: dict[int, str] = {}

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            non_empty = [cell for cell in row if cell.value is not None]
            if non_empty:
                header_row = row_idx
                headers = {cell.column: str(cell.value) for cell in row if cell.value is not None}
                break

        if header_row is None:
            report["sheet_details"].append(sheet_info)
            report["sheets_processed"] += 1
            continue

        masked_cols_in_sheet = set()

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                if cell.value is None:
                    continue

                col_name = headers.get(cell.column, f"col_{cell.column}")
                sheet_info["cells_scanned"] += 1
                report["total_cells_scanned"] += 1

                do_mask, pii_type = should_mask(col_name, cell.value)
                if do_mask:
                    cell.value = mask_value(cell.value, pii_type)
                    if highlight_masked:
                        cell.fill = MASKED_FILL

                    sheet_info["pii_found"] += 1
                    report["total_pii_masked"] += 1
                    report["pii_breakdown"][pii_type] = report["pii_breakdown"].get(pii_type, 0) + 1
                    masked_cols_in_sheet.add(col_name)

        sheet_info["columns_masked"] = sorted(masked_cols_in_sheet)
        report["sheet_details"].append(sheet_info)
        report["sheets_processed"] += 1

    try:
        wb.save(output_path)
    except Exception as e:
        report["errors"].append(f"Could not save workbook: {e}")

    return report
